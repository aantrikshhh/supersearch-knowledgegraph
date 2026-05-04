"""Loads and queries all 3 complementary graphs + accessory taxonomies.

Data sources:
- complimentary_colours_graph.xlsx (453 rows): color harmony pairings
- complimentary_product_graph.xlsx (482 rows): outfit combo definitions
- complimentary_features_graph.xlsx (140 rows): feature pairings
- bag_info.csv, shoes_info.csv, jewellery_info.csv: accessory taxonomies
"""

import csv
import openpyxl
from collections import defaultdict
from config import (
    COMP_COLOURS_PATH, COMP_PRODUCTS_PATH, COMP_FEATURES_PATH,
    BAG_INFO_PATH, SHOES_INFO_PATH, JEWELLERY_INFO_PATH,
)


class ComplementaryGraphs:
    def __init__(self):
        self.color_pairs = self._load_colours()
        self.product_combos = self._load_products()
        self.feature_pairs = self._load_features()
        self.accessories = self._load_accessories()

    def _load_colours(self):
        """Load color harmony graph. Returns dict[color] -> list of {color2, rank, context, harmony}."""
        pairs = defaultdict(list)
        wb = openpyxl.load_workbook(COMP_COLOURS_PATH, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            c1, c2, rank, context, harmony = row[0], row[1], row[2], row[3], row[4]
            pairs[str(c1).lower()].append({
                "color": str(c2).lower(),
                "rank": int(rank) if rank else 1,
                "context": str(context).lower() if context else "all",
                "harmony": str(harmony).lower() if harmony else "neutral",
            })
        wb.close()
        return dict(pairs)

    def _load_products(self):
        """Load outfit combo graph. Returns dict[combo_id] -> {name, occasion, gender, items}."""
        combos = defaultdict(lambda: {"items": []})
        wb = openpyxl.load_workbook(COMP_PRODUCTS_PATH, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            combo_id = int(row[0])
            combos[combo_id]["name"] = str(row[1]) if row[1] else ""
            combos[combo_id]["occasion_tag"] = str(row[2]).lower() if row[2] else ""
            combos[combo_id]["gender"] = str(row[3]).lower() if row[3] else "both"
            combos[combo_id]["items"].append({
                "category": str(row[4]).lower() if row[4] else "",
                "tag": str(row[5]).lower() if row[5] else "",
                "name": str(row[6]).lower() if row[6] else "",
                "rank": int(row[7]) if row[7] else 1,
            })
        wb.close()

        # Build occasion index for fast lookup
        self._occasion_index = defaultdict(list)
        for combo_id, combo in combos.items():
            tag = combo["occasion_tag"]
            if tag:
                self._occasion_index[tag].append(combo_id)
        return dict(combos)

    def _load_features(self):
        """Load feature pairing graph. Returns list of feature pair dicts."""
        pairs = []
        wb = openpyxl.load_workbook(COMP_FEATURES_PATH, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            pairs.append({
                "cat1": str(row[0]).lower(),
                "feat1_name": str(row[1]).lower(),
                "feat1_value": str(row[2]).lower(),
                "cat2": str(row[3]).lower(),
                "feat2_name": str(row[4]).lower(),
                "feat2_value": str(row[5]).lower(),
                "rank": int(row[6]) if row[6] else 1,
            })
        wb.close()
        return pairs

    def _load_accessories(self):
        """Load accessory taxonomies (bags, shoes, jewellery)."""
        result = {}
        for category, path in [("bag", BAG_INFO_PATH), ("shoes", SHOES_INFO_PATH),
                                ("jewellery", JEWELLERY_INFO_PATH)]:
            items = {}
            with open(path) as f:
                reader = csv.DictReader(f)
                for row in reader:
                    product = row.get("product", "").strip().lower()
                    types = row.get("product_type", "").strip()
                    if product:
                        items[product] = [t.strip() for t in types.split("/") if t.strip()] if types else []
            result[category] = items
        return result

    # ── Query methods ──

    def get_complementary_colors(self, color, context="all", top_n=5):
        """Get colors that pair well with the given color.

        Args:
            color: primary color (e.g., "red")
            context: one of "all", "casual", "ethnic", "formal", "party", "western"
            top_n: max results

        Returns:
            list of {color, rank, harmony} sorted by rank
        """
        color = color.lower()
        pairs = self.color_pairs.get(color, [])

        if context != "all":
            context_pairs = [p for p in pairs if p["context"] in (context, "all")]
            if context_pairs:
                pairs = context_pairs

        pairs.sort(key=lambda x: x["rank"])
        return pairs[:top_n]

    def get_outfit_combos(self, occasion_tag, gender="female"):
        """Get outfit combo definitions for an occasion.

        Args:
            occasion_tag: e.g., "casual", "sangeet", "party", "ethnic"
            gender: "male", "female", "both"

        Returns:
            list of combos, each with {name, items: [{category, tag, name, rank}]}
        """
        occasion_tag = occasion_tag.lower()
        combo_ids = self._occasion_index.get(occasion_tag, [])

        results = []
        for cid in combo_ids:
            combo = self.product_combos[cid]
            if combo["gender"] in (gender, "both"):
                results.append(combo)

        # If no exact match, try broader tags
        if not results:
            broader = {"sangeet": "ethnic", "mehendi": "ethnic", "haldi": "ethnic",
                       "wedding": "ethnic", "hindu wedding": "ethnic",
                       "muslim wedding": "ethnic", "christian wedding": "formal",
                       "engagement": "formal", "roka": "semi-formal",
                       "bachelorette": "party", "birthday party": "party",
                       "date night": "casual", "concert": "casual",
                       "office party": "semi-formal", "interview": "formal",
                       "graduation": "semi-formal", "prom": "party",
                       "corporate event": "formal", "gala": "formal"}
            broader_tag = broader.get(occasion_tag)
            if broader_tag and broader_tag != occasion_tag:
                return self.get_outfit_combos(broader_tag, gender)

        return results

    def get_complementary_features(self, category, feature_name, feature_value):
        """Get features that complement a given feature.

        Args:
            category: "clothing", "shoes", etc.
            feature_name: "fit", "neck", "sleeve", etc.
            feature_value: "baggy", "V-neck", etc.

        Returns:
            list of {category, feature_name, feature_value, rank}
        """
        category = category.lower()
        feature_name = feature_name.lower()
        feature_value = feature_value.lower()

        results = []
        for pair in self.feature_pairs:
            if (pair["cat1"] == category and pair["feat1_name"] == feature_name
                    and pair["feat1_value"] == feature_value):
                results.append({
                    "category": pair["cat2"],
                    "feature_name": pair["feat2_name"],
                    "feature_value": pair["feat2_value"],
                    "rank": pair["rank"],
                })
        results.sort(key=lambda x: x["rank"])
        return results

    def get_accessory_types(self, product_name, category):
        """Get specific sub-types for an accessory product.

        Args:
            product_name: e.g., "heel", "clutch", "maangtika"
            category: "shoes", "bag", "jewellery"

        Returns:
            list of specific type strings
        """
        cat_data = self.accessories.get(category, {})
        return cat_data.get(product_name.lower(), [])

    def get_outfit_accessories(self, occasion_tag, gender="female"):
        """Get recommended accessories (shoes, bag, jewellery) for an occasion.

        Extracts non-clothing items from outfit combos.

        Returns:
            dict with {shoes: [...], bag: [...], jewellery: [...]}
        """
        combos = self.get_outfit_combos(occasion_tag, gender)
        accessories = {"shoes": [], "bag": [], "jewellery": []}

        for combo in combos:
            for item in combo["items"]:
                cat = item["category"]
                if cat in ("shoes", "bag", "jewellery"):
                    if item["name"] not in [a["name"] for a in accessories.get(cat, [])]:
                        accessories.setdefault(cat, []).append({
                            "name": item["name"],
                            "rank": item["rank"],
                        })

        for cat in accessories:
            accessories[cat].sort(key=lambda x: x["rank"])

        return accessories
