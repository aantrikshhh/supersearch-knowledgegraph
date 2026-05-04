import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

random.seed(42)

# Entity values from Pre-Graph data
entities = {
    "Place": ["temple", "mosque", "park", "office", "beach", "mountains", "restaurant", "club", "gym", "mall", "cafe", "museum", "library", "stadium", "theater", "church", "garden", "supermarket", "hospital", "airport"],
    "Profession": ["teacher", "nurse", "driver", "engineer", "doctor", "artist", "chef", "pilot", "lawyer", "writer", "designer", "developer", "accountant", "manager", "photographer"],
    "Activity": ["running", "hiking", "swimming", "cycling", "dancing", "yoga", "cooking", "traveling", "fishing", "skiing", "vacation"],
    "Occasion": ["christian wedding", "bachelorette", "baby shower", "birthday party", "graduation", "anniversary", "corporate event", "date night", "festival", "concert", "picnic", "gala", "funeral", "reunion", "interview", "office party", "housewarming", "engagement", "prom", "retirement party", "farewell party", "haldi", "sangeet", "mehendi", "roka", "Indian wedding", "muslim wedding", "griha pravesh", "rice ceremony"],
    "Month": ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"],
    "Weather": ["summer", "winter", "rainy", "sunny", "cloudy", "snowy", "windy", "humid", "dry", "stormy"],
    "Relation": ["cousin", "mom", "sister", "niece", "friend", "brother", "dad", "uncle", "aunt", "nephew", "grandfather", "grandmother", "colleague", "boss"],
    "Health": ["knee pain", "sweating", "hunch back", "back pain", "asthma", "diabetes", "arthritis", "flat feet", "sensitive skin", "allergies", "migraines", "poor circulation", "bad posture"],
    "BodyType": ["flat chested", "pear shaped", "V-shaped", "hourglass", "apple shaped", "athletic", "petite", "tall", "plus size", "slim", "broad shoulders", "short torso", "long legs"],
    "Time": ["morning", "evening", "afternoon", "night", "midnight", "sunrise", "sunset", "dawn", "dusk", "noon", "early morning", "late night"],
    "Budget": ["cheap", "expensive", "affordable", "luxury", "mid-range", "budget-friendly", "high-end", "low-cost", "premium", "economical", "extravagant", "moderate", "reasonable"],
    "AgeGroup": ["teenager", "adult", "senior", "child", "middle-aged", "young adult", "elderly", "tween", "mature", "youth"],
    "Religion": ["Christianity", "Islam", "Hinduism", "Buddhism", "Sikhism", "Judaism"],
    "Complexion": ["fair", "wheatish", "dark", "brown", "olive"],
    "Event": ["Diwali", "Christmas", "Coachella", "Oktoberfest", "Rio Carnival", "Holi", "Ganesh Chaturthi", "Puja", "Pongal", "Onam", "Disneyland", "Navratri", "Bihu", "Eid", "Lohri"],
    "Location": ["India", "Mumbai", "Delhi", "Bangalore", "Jaipur", "Goa", "Dubai", "London", "New York", "Paris", "Hyderabad", "Chennai", "Kolkata", "Lucknow", "Udaipur"],
}

# Brand profiles
brands = {
    "Masaba": {
        "description": "Bold prints, contemporary Indian fusion, sarees, kurtas, dresses",
        "occasions": ["Indian wedding", "sangeet", "mehendi", "haldi", "engagement", "date night", "birthday party", "festival", "concert", "bachelorette", "gala", "anniversary", "office party", "farewell party", "roka", "baby shower", "graduation", "prom"],
        "events": ["Diwali", "Holi", "Navratri", "Ganesh Chaturthi", "Pongal", "Onam", "Bihu", "Christmas", "Coachella"],
        "places": ["restaurant", "club", "cafe", "mall", "office", "beach", "garden", "theater", "museum", "airport"],
        "budgets": ["mid-range", "premium", "high-end", "luxury", "affordable"],
        "activities": ["dancing", "traveling", "vacation"],
        "templates": [
            "What should I wear to my {Relation}'s {Occasion}?",
            "Show me something bold for {Occasion} in {Month}",
            "I need a printed outfit for {Event}",
            "What to wear to a {Place} in {Weather} weather?",
            "Looking for a fusion outfit for {Occasion} in {Location}",
            "Suggest a {Budget} outfit for my {Relation}'s {Occasion}",
            "I'm {BodyType} — what works for {Occasion}?",
            "Outfit ideas for {Event} in {Location}",
            "What can a {Profession} wear to a {Occasion}?",
            "Need a printed kurta set for {Occasion} in {Month}",
            "Show me something for a {AgeGroup} going to {Occasion}",
            "What to wear to {Event} if I'm on a {Budget} budget?",
            "I have {Health} — suggest comfortable ethnic wear for {Occasion}",
            "What should I wear to a {Place} in {Month}?",
            "Looking for {Weather} wedding guest outfits",
            "Bold saree options for {Occasion}",
            "Suggest a dress for {Occasion} in the {Time}",
            "What to wear to {Place} if I practice {Religion}?",
            "Outfit for {Activity} in {Location} during {Weather}",
            "Need a {Budget} outfit for my {AgeGroup} {Relation} for {Occasion}",
        ],
    },
    "Indya": {
        "description": "Ethnic fusion, kurtas, palazzos, co-ords, young women's Indian wear",
        "occasions": ["Indian wedding", "sangeet", "mehendi", "haldi", "festival", "office party", "date night", "engagement", "birthday party", "college fest", "roka", "farewell party", "bachelorette", "baby shower", "housewarming", "griha pravesh", "prom", "graduation", "anniversary", "concert"],
        "events": ["Diwali", "Navratri", "Holi", "Ganesh Chaturthi", "Puja", "Pongal", "Onam", "Bihu", "Eid", "Lohri", "Christmas"],
        "places": ["office", "temple", "mall", "cafe", "restaurant", "garden", "park", "museum", "airport", "church"],
        "budgets": ["affordable", "budget-friendly", "mid-range", "economical", "moderate", "reasonable", "low-cost"],
        "activities": ["traveling", "vacation", "yoga", "dancing"],
        "templates": [
            "Kurta set for {Occasion} in {Month}",
            "What ethnic wear for {Occasion} in {Weather} weather?",
            "Affordable outfit for my {Relation}'s {Occasion}",
            "Fusion wear for {Event} in {Location}",
            "I'm a {Profession} — what to wear to {Occasion}?",
            "Suggest a {Budget} kurta for {Occasion}",
            "Co-ord set ideas for {Occasion} in the {Time}",
            "What to wear to a {Place} in {Month}?",
            "Need palazzo set for {AgeGroup} for {Occasion}",
            "Ethnic outfit for {Activity} in {Location}",
            "What should I wear to {Event}?",
            "Comfortable ethnic wear — I have {Health}",
            "I'm {BodyType}, suggest a kurta set for {Occasion}",
            "Outfit for {Occasion} on a {Budget} budget",
            "What to wear to {Place} during {Weather}?",
            "Show me festive kurtas for {Event} in {Location}",
            "Need something for my {Relation}'s {Occasion} in {Month}",
            "What's appropriate for a {Religion} {Occasion}?",
            "Suggest a {Budget} outfit for {Event}",
            "Ethnic fusion look for {Place} in {Weather} weather",
        ],
    },
    "Rothys": {
        "description": "Sustainable shoes (flats, sneakers, loafers), bags, accessories",
        "occasions": ["corporate event", "interview", "office party", "date night", "graduation", "birthday party", "concert", "reunion", "farewell party", "gala", "picnic", "housewarming"],
        "events": ["Coachella", "Disneyland", "Christmas", "Oktoberfest"],
        "places": ["office", "restaurant", "mall", "cafe", "museum", "library", "airport", "park", "garden", "theater", "gym"],
        "budgets": ["premium", "mid-range", "high-end", "luxury", "moderate"],
        "activities": ["traveling", "hiking", "vacation", "cycling", "yoga", "dancing"],
        "templates": [
            "Comfortable shoes for {Activity} all day",
            "What flats work for a {Profession} at {Place}?",
            "Need shoes for {Occasion} that are comfortable with {Health}",
            "Best flats for {Weather} weather in {Location}",
            "Shoes for {Activity} in the {Time}",
            "What to wear on my feet to {Occasion}?",
            "Suggest {Budget} sustainable shoes for {Place}",
            "I'm a {Profession} — need shoes for {Place} in {Month}",
            "Comfortable flats for a {AgeGroup} going to {Occasion}",
            "Shoes for {Event} in {Location}",
            "What shoes for {Place} during {Weather}?",
            "Need loafers for {Occasion} in {Month}",
            "Sneakers for {Activity} with {Health}",
            "Bags for {Occasion} in {Weather} weather",
            "Show me shoes that work for a {BodyType} person doing {Activity}",
            "Sustainable shoes for {Place} in {Month}",
            "Walking shoes for {Activity} in {Location}",
            "Slip-ons for {Occasion} at {Place}",
            "What shoes for a {Profession} attending {Occasion}?",
            "Comfortable shoes for {Event} — I have {Health}",
        ],
    },
    "Mud": {
        "description": "Casual contemporary, denim, sustainable fashion, everyday wear",
        "occasions": ["date night", "birthday party", "concert", "picnic", "reunion", "farewell party", "office party", "bachelorette", "housewarming"],
        "events": ["Coachella", "Christmas", "Oktoberfest", "Disneyland", "Holi"],
        "places": ["cafe", "mall", "park", "restaurant", "club", "beach", "mountains", "stadium", "garden", "office", "airport"],
        "budgets": ["affordable", "mid-range", "budget-friendly", "moderate", "reasonable", "economical"],
        "activities": ["hiking", "cycling", "traveling", "vacation", "fishing", "camping"],
        "templates": [
            "Casual outfit for {Place} in {Weather} weather",
            "Jeans for {Activity} in {Month}",
            "What to wear to a {Occasion} casually?",
            "Suggest {Budget} denim for {Activity}",
            "Comfortable outfit for {Place} in {Month}",
            "I'm {BodyType} — what jeans work for me?",
            "Casual wear for {Event} in {Location}",
            "Weekend outfit for {Place} during {Weather}",
            "What should a {Profession} wear casually to {Place}?",
            "Sustainable outfit for {Activity} in the {Time}",
            "Denim outfit for {Occasion} in {Weather}",
            "Need {Budget} casual wear for {AgeGroup}",
            "What to wear to {Place} with {Health}?",
            "Casual outfit for my {Relation}'s {Occasion}",
            "Outdoor wear for {Activity} in {Location} in {Month}",
            "Everyday outfit for {Place} in {Weather} weather",
            "Comfortable jeans for {Activity} with {Health}",
            "Casual look for {Event}",
            "Show me something for {Occasion} on a {Budget} budget",
            "What to wear to {Place} in {Month} for a {AgeGroup}?",
        ],
    },
    "Kalki": {
        "description": "Luxury bridal, lehengas, sarees, gowns, anarkalis, heavy ethnic",
        "occasions": ["Indian wedding", "sangeet", "mehendi", "haldi", "engagement", "roka", "reception", "christian wedding", "muslim wedding", "anniversary", "gala", "festival", "baby shower", "griha pravesh", "rice ceremony", "graduation", "prom", "bachelorette", "farewell party", "retirement party"],
        "events": ["Diwali", "Navratri", "Ganesh Chaturthi", "Puja", "Pongal", "Onam", "Bihu", "Eid", "Lohri", "Christmas", "Holi"],
        "places": ["temple", "mosque", "church", "restaurant", "theater", "garden", "mall", "museum"],
        "budgets": ["luxury", "premium", "high-end", "extravagant", "expensive"],
        "activities": ["dancing"],
        "templates": [
            "Lehenga for {Occasion} in {Month}",
            "What to wear to my {Relation}'s {Occasion} in {Location}?",
            "Bridal outfit for {Occasion} in {Weather} weather",
            "{Budget} lehenga for {Occasion}",
            "Saree for {Event} in {Location}",
            "I'm {BodyType} — suggest a lehenga for {Occasion}",
            "Gown for {Occasion} in the {Time}",
            "What should a {AgeGroup} wear to {Occasion}?",
            "Heavy ethnic outfit for {Occasion} in {Month}",
            "Anarkali for {Event}",
            "Outfit for {Occasion} at a {Place} in {Location}",
            "Designer saree for my {Relation}'s {Occasion}",
            "Sharara set for {Occasion} in {Month}",
            "What to wear to a {Religion} {Occasion}?",
            "Suggest {Budget} bridal wear for {Occasion} in {Location}",
            "Embroidered outfit for {Event} in {Month}",
            "Need a {Budget} outfit for {Occasion} — I'm {BodyType}",
            "Wedding guest lehenga for {Occasion} in {Weather}",
            "Show me festive sarees for {Event}",
            "What to wear to {Occasion} if I have {Health}?",
        ],
    },
    "Aza": {
        "description": "Multi-designer luxury, bridal couture, high fashion Indian wear",
        "occasions": ["Indian wedding", "sangeet", "mehendi", "haldi", "engagement", "roka", "christian wedding", "muslim wedding", "gala", "anniversary", "reception", "bachelorette", "baby shower", "graduation", "prom", "concert", "corporate event", "date night", "farewell party"],
        "events": ["Diwali", "Navratri", "Ganesh Chaturthi", "Puja", "Christmas", "Eid", "Holi", "Onam", "Bihu", "Pongal"],
        "places": ["restaurant", "theater", "museum", "garden", "club", "temple", "church", "mosque", "mall"],
        "budgets": ["luxury", "premium", "high-end", "extravagant", "expensive"],
        "activities": ["dancing", "traveling", "vacation"],
        "templates": [
            "Designer outfit for {Occasion} in {Location}",
            "What designer wear for my {Relation}'s {Occasion}?",
            "Luxury lehenga for {Occasion} in {Month}",
            "Couture saree for {Event}",
            "I'm {BodyType} — suggest designer wear for {Occasion}",
            "{Budget} designer outfit for {Occasion} in {Weather}",
            "What to wear to a {Religion} {Occasion} in {Location}?",
            "Designer gown for {Occasion} in the {Time}",
            "Show me {Budget} options for {Event} in {Location}",
            "Bridal couture for {Occasion} in {Month}",
            "What should a {AgeGroup} wear to {Occasion}?",
            "Designer outfit for {Place} in {Weather} weather",
            "Suggest a {Budget} outfit for my {Relation}'s {Occasion}",
            "Multi-designer look for {Event}",
            "Heavy embroidered outfit for {Occasion}",
            "What to wear to {Occasion} at a {Place} in {Location}?",
            "I'm a {Profession} attending {Occasion} — need something {Budget}",
            "Luxury saree for {Event} in {Month}",
            "What's appropriate for {Occasion} if I have {Health}?",
            "Designer wear for {AgeGroup} at {Occasion}",
        ],
    },
    "Shoppers Stop": {
        "description": "Department store — casual, formal, ethnic, western, beauty, all segments",
        "occasions": ["Indian wedding", "sangeet", "birthday party", "corporate event", "date night", "interview", "office party", "festival", "concert", "graduation", "farewell party", "reunion", "housewarming", "engagement", "baby shower", "bachelorette", "prom", "retirement party", "picnic", "funeral", "anniversary", "mehendi", "haldi"],
        "events": ["Diwali", "Christmas", "Holi", "Navratri", "Ganesh Chaturthi", "Puja", "Pongal", "Onam", "Eid", "Bihu", "Lohri", "Coachella", "Disneyland"],
        "places": ["office", "mall", "restaurant", "cafe", "park", "beach", "temple", "church", "mosque", "gym", "airport", "club", "garden", "museum", "theater", "library", "stadium", "hospital", "supermarket", "mountains"],
        "budgets": ["affordable", "budget-friendly", "mid-range", "economical", "moderate", "reasonable", "cheap", "low-cost", "premium", "high-end"],
        "activities": ["running", "hiking", "swimming", "cycling", "dancing", "yoga", "traveling", "vacation", "skiing", "fishing"],
        "templates": [
            "What to wear to {Occasion} in {Month}?",
            "Outfit for {Place} in {Weather} weather",
            "I'm a {Profession} — need {Budget} work clothes",
            "Suggest clothes for {Activity} in {Location}",
            "What should a {AgeGroup} wear to {Occasion}?",
            "{Budget} outfit for {Event}",
            "I'm {BodyType} — what to wear to {Occasion}?",
            "Formal wear for {Occasion} in the {Time}",
            "Outfit for my {Relation}'s {Occasion} in {Month}",
            "What to wear to {Place} during {Weather}?",
            "Comfortable outfit for {Activity} with {Health}",
            "Casual wear for {Place} in {Month}",
            "Show me {Budget} options for {Occasion}",
            "What is appropriate for a {Religion} person at {Occasion}?",
            "Kids outfit for {Event} in {Location}",
            "Need clothes for my {Relation} who is a {AgeGroup}",
            "What should a {Profession} wear to {Place} in {Month}?",
            "Party wear for {Occasion} in {Weather}",
            "Suggest an outfit for {Event} if I am on a {Budget} budget",
            "What can I wear to a {Place} in {Location}?",
        ],
    },
    "Reliance": {
        "description": "Mass market (Ajio/Trends) — all categories, budget to mid-range",
        "occasions": ["Indian wedding", "birthday party", "corporate event", "date night", "festival", "concert", "office party", "interview", "graduation", "farewell party", "reunion", "picnic", "housewarming", "engagement", "baby shower", "sangeet", "mehendi", "haldi", "bachelorette", "prom", "anniversary", "funeral", "retirement party"],
        "events": ["Diwali", "Christmas", "Holi", "Navratri", "Ganesh Chaturthi", "Puja", "Pongal", "Onam", "Eid", "Bihu", "Lohri", "Coachella", "Disneyland"],
        "places": ["office", "mall", "restaurant", "cafe", "park", "beach", "temple", "church", "mosque", "gym", "airport", "club", "garden", "museum", "theater", "library", "stadium", "mountains"],
        "budgets": ["cheap", "affordable", "budget-friendly", "economical", "low-cost", "moderate", "reasonable", "mid-range"],
        "activities": ["running", "hiking", "swimming", "cycling", "dancing", "yoga", "traveling", "vacation", "fishing"],
        "templates": [
            "Cheap outfit for {Occasion}",
            "What to wear to {Place} on a {Budget} budget?",
            "Affordable {Weather} wear in {Location}",
            "Budget outfit for my {Relation}'s {Occasion}",
            "I'm a {Profession} — need {Budget} clothes for {Place}",
            "{Budget} ethnic wear for {Event}",
            "Suggest clothes for {Activity} under budget",
            "I'm {BodyType} — {Budget} options for {Occasion}?",
            "What should a {AgeGroup} wear to {Occasion} in {Month}?",
            "Casual {Budget} outfit for {Place} in {Weather}",
            "Comfortable clothes for {Activity} with {Health}",
            "What to wear to {Event} in {Location}?",
            "{Budget} festive wear for {Event} in {Month}",
            "Need {Budget} clothes for my {Relation} for {Occasion}",
            "What can I wear to {Place} during {Weather} weather?",
            "Office wear for {Profession} in {Month}",
            "Outfit ideas for {Occasion} in the {Time}",
            "What is appropriate for a {Religion} event at {Place}?",
            "Show me {Budget} options for {AgeGroup} at {Occasion}",
            "Everyday outfit for {Place} in {Month}",
        ],
    },
    "Lifestyle": {
        "description": "Mid-range department store — western, ethnic, casual, formal",
        "occasions": ["birthday party", "corporate event", "date night", "interview", "office party", "festival", "concert", "graduation", "farewell party", "reunion", "Indian wedding", "sangeet", "engagement", "housewarming", "baby shower", "bachelorette", "prom", "retirement party", "picnic", "anniversary", "mehendi"],
        "events": ["Diwali", "Christmas", "Holi", "Navratri", "Ganesh Chaturthi", "Puja", "Pongal", "Onam", "Eid", "Coachella", "Disneyland"],
        "places": ["office", "mall", "restaurant", "cafe", "park", "beach", "temple", "gym", "airport", "club", "garden", "museum", "theater", "mountains"],
        "budgets": ["affordable", "mid-range", "moderate", "reasonable", "budget-friendly", "premium", "economical"],
        "activities": ["running", "hiking", "cycling", "dancing", "yoga", "traveling", "vacation", "swimming"],
        "templates": [
            "What to wear to {Occasion} in {Month}?",
            "Outfit for {Place} in {Weather} weather",
            "I'm a {Profession} — suggest work wear for {Month}",
            "Clothes for {Activity} in {Location}",
            "{Budget} outfit for {Event}",
            "I'm {BodyType} — what looks good for {Occasion}?",
            "What should a {AgeGroup} wear to {Occasion}?",
            "Suggest {Weather} outfits for {Place}",
            "Party wear for {Occasion} in the {Time}",
            "Outfit for my {Relation}'s {Occasion}",
            "Casual wear for {Place} in {Month}",
            "Comfortable outfit for {Activity} — I have {Health}",
            "What to wear to {Event} in {Location}?",
            "Show me {Budget} options for {Occasion} in {Weather}",
            "Need clothes for my {Relation} who is a {AgeGroup}",
            "What should a {Profession} wear to {Place}?",
            "Festive outfit for {Event} in {Month}",
            "What is appropriate for {Religion} {Occasion}?",
            "Suggest an outfit for {Occasion} on a {Budget} budget",
            "Everyday outfit for {Place} during {Weather}",
        ],
    },
}


def extract_intents(template, values):
    intents = {}
    for key, val in values.items():
        placeholder = "{" + key + "}"
        if placeholder in template:
            intents[key.lower()] = val
    return intents


def generate_query(template, brand_profile):
    values = {}
    all_keys = ["Place", "Profession", "Activity", "Occasion", "Month", "Weather",
                 "Relation", "Health", "BodyType", "Time", "Budget", "AgeGroup",
                 "Religion", "Complexion", "Event", "Location"]

    for key in all_keys:
        if "{" + key + "}" in template:
            if key == "Place" and "places" in brand_profile:
                values[key] = random.choice(brand_profile["places"])
            elif key == "Occasion" and "occasions" in brand_profile:
                values[key] = random.choice(brand_profile["occasions"])
            elif key == "Event" and "events" in brand_profile:
                values[key] = random.choice(brand_profile["events"])
            elif key == "Budget" and "budgets" in brand_profile:
                values[key] = random.choice(brand_profile["budgets"])
            elif key == "Activity" and "activities" in brand_profile:
                values[key] = random.choice(brand_profile["activities"])
            elif key in entities:
                values[key] = random.choice(entities[key])
            else:
                values[key] = ""

    query = template
    for key, val in values.items():
        query = query.replace("{" + key + "}", val)

    # Fix grammar: "a" before vowel sounds -> "an"
    import re
    query = re.sub(r'\ba ([aeiouAEIOU])', r'an \1', query)

    intents = extract_intents(template, values)
    return query, intents


wb = openpyxl.Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

QUERIES_PER_BRAND = 100

for brand_name, profile in brands.items():
    ws = wb.create_sheet(title=brand_name[:31])

    headers = ["#", "Query", "Intents", "Entity", "Entity Value", "Tag"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    seen = set()
    row_num = 2
    attempts = 0

    while len(seen) < QUERIES_PER_BRAND and attempts < QUERIES_PER_BRAND * 5:
        attempts += 1
        template = random.choice(profile["templates"])
        query, intents = generate_query(template, profile)

        if query in seen:
            continue
        seen.add(query)

        intent_str = "{ " + ", ".join(f"{k}: {v}" for k, v in intents.items()) + " }"

        # Break intents into entity/value pairs for the graph lookup columns
        intent_items = list(intents.items())
        if intent_items:
            first_entity, first_value = intent_items[0]
            tag = ""
            if first_entity in ("place", "occasion", "activity", "event"):
                tag = "product"
            elif first_entity == "weather":
                tag = "pattern"
            elif first_entity == "budget":
                tag = "product"
        else:
            first_entity, first_value, tag = "", "", ""

        ws.cell(row=row_num, column=1, value=row_num - 1).border = thin_border
        ws.cell(row=row_num, column=2, value=query).border = thin_border
        ws.cell(row=row_num, column=3, value=intent_str).border = thin_border
        ws.cell(row=row_num, column=4, value=first_entity).border = thin_border
        ws.cell(row=row_num, column=5, value=first_value).border = thin_border
        ws.cell(row=row_num, column=6, value=tag).border = thin_border

        row_num += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 12

# Summary sheet
ws_summary = wb.create_sheet(title="Summary", index=0)
summary_headers = ["Brand", "Total Queries", "Focus"]
for col, h in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for i, (brand_name, profile) in enumerate(brands.items(), 2):
    ws_summary.cell(row=i, column=1, value=brand_name).border = thin_border
    ws_summary.cell(row=i, column=2, value=QUERIES_PER_BRAND).border = thin_border
    ws_summary.cell(row=i, column=3, value=profile["description"]).border = thin_border

ws_summary.column_dimensions["A"].width = 18
ws_summary.column_dimensions["B"].width = 15
ws_summary.column_dimensions["C"].width = 65

output_path = "Brand_Queries.xlsx"
wb.save(output_path)
print(f"Generated {output_path} with {len(brands)} brand sheets, {QUERIES_PER_BRAND} queries each")
print(f"Total queries: {len(brands) * QUERIES_PER_BRAND}")

# Print a few samples per brand
for brand_name, profile in brands.items():
    print(f"\n--- {brand_name} (samples) ---")
    random.seed(brand_name)
    templates = random.sample(profile["templates"], min(5, len(profile["templates"])))
    for t in templates:
        q, intents = generate_query(t, profile)
        intent_str = "{ " + ", ".join(f"{k}: {v}" for k, v in intents.items()) + " }"
        print(f"  Q: {q}")
        print(f"     {intent_str}")
